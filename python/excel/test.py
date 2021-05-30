from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook("./test.xlsx")

ws = wb.active	# give active worksheet

print(ws)

print(wb.sheetnames)

ws = wb["Sheet1"]

print(ws["A1"].value)
print(ws["A2"].value)

ws["A1"].value = "AAAA"


wb.create_sheet("TestSheet")

ws = wb["TestSheet"]

ws["A1"].value = "AAAA"

wb.save("test2.xlsx")

###################

wb = Workbook()
ws = wb.active
ws.title = "Data"

for i in range(1, 100):
	ws[f"A{i}"].value = i
	ws[f"B{i}"].value = i * i

ws.append(["This", "is", "a", "test"])

for row in range(1, 21):
	for col in range(3, 30):
		char = get_column_letter(col)
		s = ws[char + str(row)]
		s.value = row * col
		if (row * col) % 2 == 1:
			s.font = Font(bold = True, color = "00AA0000")

	charl = get_column_letter(30)
	print(f"=SUM(C{row}:{char}{row})")
	ws[charl + str(row)] = f"=SUM(C{row}:{char}{row})"


# ws.insert_rows(10)
# ws.insert_cols(10)

wb.save("test3.xlsx")
